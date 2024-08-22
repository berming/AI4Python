import openpyxl
from openpyxl.styles import PatternFill
import calendar
from datetime import datetime, timedelta

# 创建一个新的Excel工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GitHub Contribution Calendar"

# 示例数据：每一天的贡献数
# 可以根据需要修改此数据
contribution_data = {
    (2023, 1, 1): 0,
    (2023, 1, 2): 3,
    (2023, 1, 3): 5,
    (2023, 1, 4): 2,
    (2023, 1, 5): 4,
    (2023, 1, 6): 1,
    (2023, 1, 7): 0,
    # 添加更多数据...
}

# 定义填充颜色
colors = [
    "FFFFFF",  # 0 contributions: 白色
    "D6E685",  # 1-2 contributions: 浅绿色
    "8CC665",  # 3-4 contributions: 中绿色
    "44A340",  # 5-6 contributions: 深绿色
    "1E6823",  # 7+ contributions: 最深绿色
]

def get_fill_color(count):
    if count == 0:
        return colors[0]
    elif 1 <= count <= 2:
        return colors[1]
    elif 3 <= count <= 4:
        return colors[2]
    elif 5 <= count <= 6:
        return colors[3]
    else:
        return colors[4]

# 创建日历布局
start_date = datetime(2023, 1, 1)
end_date = datetime(2023, 12, 31)

# 星期天到星期六的列号
start_col = 2

# 填写月份标题
for month in range(1, 13):
    month_name = calendar.month_name[month]
    ws.cell(row=1, column=start_col + (month - 1) * 8, value=month_name)

# 填写每一天
current_date = start_date
while current_date <= end_date:
    week_day = current_date.weekday()  # Monday = 0, Sunday = 6
    col = start_col + (current_date.month - 1) * 8 + week_day
    week_of_month = (current_date.day - 1) // 7
    row = 2 + (week_of_month * 7) + (current_date.isocalendar()[1] % 7)

    count = contribution_data.get((current_date.year, current_date.month, current_date.day), 0)
    fill_color = get_fill_color(count)
    cell = ws.cell(row=row, column=col, value=current_date.day)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    current_date += timedelta(days=1)

# 保存工作簿
wb.save("github_contribution_calendar.xlsx")
